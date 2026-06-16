import openpyxl
from pathlib import Path
from typing import List, Dict, Any
from extractor.logger import logger

def export_to_excel(records: List[Dict[str, Any]], template_path: str, output_path: str) -> None:
    """
    Loads the Excel template, clears the dummy template data on Row 2,
    maps voter records to their target columns dynamically by matching header names,
    saves the workbook to the specified output path.
    """
    logger.info(f"Exporting {len(records)} records to Excel template at: {template_path}")
    
    t_path = Path(template_path)
    if not t_path.exists():
        logger.error(f"Excel template file not found at: {template_path}")
        raise FileNotFoundError(f"Excel template not found at {template_path}")
        
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    # 1. Map headers to their 1-indexed column indices
    headers = {}
    for col in range(1, sheet.max_column + 1):
        cell_val = sheet.cell(row=1, column=col).value
        if cell_val:
            headers[str(cell_val).strip()] = col
            
    logger.info(f"Excel headers found: {list(headers.keys())}")
    
    # Mapping between record dictionary keys and the Excel template column header names
    key_mapping = {
        "ac_no": "AC_NO",
        "ac_name": "AC NAME",
        "section_no": "SECTION",
        "booth": "Booth ",
        "village_area": "Villages/Area",
        "booth_no": "BOOTH No. ",
        "sl_no": "SL",
        "epic_no": "EPIC",
        "house_no": "house_number",
        "name": "NAME",
        "relation_type": "R_TYPE",
        "relation_name": "R_NAME",
        "age": "AGE",
        "gender": "SEX",
        "booth_name": "BOOTH NAME"
    }
    
    # 2. Clear dummy rows starting from Row 2
    max_row = sheet.max_row
    if max_row >= 2:
        sheet.delete_rows(2, max_row)
        logger.info(f"Deleted {max_row - 1} dummy/existing rows from the template worksheet.")
        
    # 3. Write records row by row
    current_row = 2
    for record in records:
        for key, header_name in key_mapping.items():
            col_idx = headers.get(header_name.strip())
            if col_idx:
                sheet.cell(row=current_row, column=col_idx, value=record.get(key, ""))
        current_row += 1
        
    # Ensure the parent directory of output_path exists
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb.save(out_path)
    logger.info(f"Successfully generated Excel export: {out_path} ({len(records)} rows)")
